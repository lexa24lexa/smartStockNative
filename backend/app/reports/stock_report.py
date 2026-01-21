from datetime import date
from io import BytesIO
from typing import List
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_stock_pdf_report(stock_data: List[dict], store_name: str, report_date: date):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18,
                                 textColor=colors.HexColor('#1a1a1a'), spaceAfter=30, alignment=1)
    elements.append(Paragraph("Daily Stock Report", title_style))
    elements.append(Spacer(1, 0.2*inch))

    info_data = [
        ['Store:', store_name],
        ['Report Date:', report_date.strftime('%Y-%m-%d')],
        ['Total Items:', str(len(stock_data))],
        ['Total Quantity:', str(sum(item['quantity'] for item in stock_data))]
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([('FONTNAME', (0,0),(0,-1), 'Helvetica-Bold'),
                                    ('FONTNAME', (1,0),(1,-1), 'Helvetica'),
                                    ('FONTSIZE', (0,0), (-1,-1), 11),
                                    ('BOTTOMPADDING', (0,0), (-1,-1), 12),
                                    ('TOPPADDING', (0,0), (-1,-1), 12)]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    if stock_data:
        table_data = [['Product', 'Batch Code', 'Expiration Date', 'Quantity']]
        for item in stock_data:
            exp_date = item['expiration_date'].strftime('%Y-%m-%d') if item['expiration_date'] else 'N/A'
            table_data.append([item['product_name'], item['batch_code'], exp_date, str(item['quantity'])])

        stock_table = Table(table_data, colWidths=[2.5*inch, 2*inch, 1.5*inch, 1*inch])
        stock_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey]),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),
        ]))
        elements.append(stock_table)
    else:
        elements.append(Paragraph("No stock items found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_stock_excel_report(stock_data: List[dict], store_name: str, report_date: date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Stock Report"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)

    ws['A1'] = "Daily Stock Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = "Store:"; ws['B3'] = store_name
    ws['A4'] = "Report Date:"; ws['B4'] = report_date.strftime('%Y-%m-%d')
    ws['A5'] = "Total Items:"; ws['B5'] = len(stock_data)
    ws['A6'] = "Total Quantity:"; ws['B6'] = sum(item['quantity'] for item in stock_data)

    headers = ['Product', 'Batch Code', 'Expiration Date', 'Quantity']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(stock_data, start=9):
        ws.cell(row=row_idx, column=1).value = item['product_name']
        ws.cell(row=row_idx, column=2).value = item['batch_code']
        ws.cell(row=row_idx, column=3).value = item['expiration_date'].strftime('%Y-%m-%d') if item['expiration_date'] else 'N/A'
        ws.cell(row=row_idx, column=4).value = item['quantity']
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

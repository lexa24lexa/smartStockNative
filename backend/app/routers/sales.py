from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from typing import List

from .. import models, schemas, database
from ..services.stock_service import StockService, FIFOService

router = APIRouter(prefix="/sales", tags=["Sales"])

# Get all sales
@router.get("", response_model=List[schemas.SaleResponse])
def get_sales(db: Session = Depends(database.get_db)):
    sales = db.query(models.Sale).order_by(models.Sale.date.desc()).all()
    result = []
    for sale in sales:
        sale_lines = []
        for line in sale.sale_lines:
            batch = line.batch
            product = batch.product
            product_response = schemas.ProductResponse(
                product_id=product.product_id,
                name=product.name,
                unit_price=product.unit_price,
                supplier_id=product.supplier_id,
                category_id=product.category_id,
                quantity=None,
                facing=None
            )
            batch_response = schemas.BatchResponse(
                batch_id=batch.batch_id,
                product_id=batch.product_id,
                batch_code=batch.batch_code,
                expiration_date=batch.expiration_date,
                product=product_response
            )
            sale_lines.append(schemas.SaleLineResponse(
                line_id=line.line_id,
                batch_id=line.batch_id,
                quantity=line.quantity,
                subtotal=line.subtotal,
                batch=batch_response
            ))
        result.append(schemas.SaleResponse(
            sale_id=sale.sale_id,
            store_id=sale.store_id,
            total_amount=sale.total_amount,
            date=sale.date,
            lines=sale_lines
        ))
    return result

# Create sale with FIFO check
@router.post("/fifo", response_model=schemas.SaleResponse)
def create_sale_fifo(sale_input: schemas.SaleCreate, db: Session = Depends(database.get_db)):
    store = StockService.get_store(db, sale_input.store_id)
    new_sale = models.Sale(store_id=store.store_id, total_amount=0)
    db.add(new_sale)
    db.commit()
    db.refresh(new_sale)

    total_amount = 0
    sale_lines_instances = []

    for line in sale_input.lines:
        if line.quantity <= 0:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Quantity must be > 0 for batch {line.batch_id}")

        batch = db.query(models.Batch).filter(models.Batch.batch_id == line.batch_id, models.Batch.is_active.is_(True)).first()
        if not batch:
            db.rollback()
            raise HTTPException(status_code=404, detail=f"Batch {line.batch_id} not found or inactive")

        product = db.query(models.Product).filter(models.Product.product_id == batch.product_id, models.Product.is_active.is_(True)).first()
        if not product:
            db.rollback()
            raise HTTPException(status_code=404, detail=f"Product for batch {line.batch_id} not found or inactive")

        fifo_check = FIFOService.check_fifo_violation(db, store.store_id, product.product_id, line.batch_id)
        if fifo_check['is_violation']:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"FIFO violation: expected batch {fifo_check['expected_batch_code']} for product {product.name}")

        stock_record = db.query(models.Stock).filter(models.Stock.store_id == store.store_id, models.Stock.batch_id == line.batch_id).first()
        if not stock_record or stock_record.quantity < line.quantity:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Insufficient stock for batch {line.batch_id}")

        stock_record.quantity -= line.quantity
        subtotal = float(product.unit_price) * line.quantity
        if line.subtotal and abs(line.subtotal - subtotal) > 0.01:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Invalid subtotal for batch {line.batch_id}")

        sale_line = models.SaleLine(sale_id=new_sale.sale_id, batch_id=line.batch_id, quantity=line.quantity, subtotal=subtotal)
        db.add(sale_line)
        sale_lines_instances.append(sale_line)
        total_amount += subtotal

    new_sale.total_amount = total_amount
    db.commit()
    db.refresh(new_sale)

    return schemas.SaleResponse(
        sale_id=new_sale.sale_id,
        store_id=new_sale.store_id,
        total_amount=new_sale.total_amount,
        date=new_sale.date,
        lines=sale_lines_instances
    )

# Get single sale by ID
@router.get("/{sale_id}", response_model=schemas.SaleResponse)
def get_sale(sale_id: int, db: Session = Depends(database.get_db)):
    sale = db.query(models.Sale).filter(models.Sale.sale_id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail=f"Sale {sale_id} not found")
    return schemas.SaleResponse(
        sale_id=sale.sale_id,
        store_id=sale.store_id,
        total_amount=sale.total_amount,
        date=sale.date,
        lines=sale.sale_lines
    )

# Update sale with FIFO check
@router.put("/{sale_id}", response_model=schemas.SaleResponse)
def update_sale_fifo(sale_id: int, sale_input: schemas.SaleCreate, db: Session = Depends(database.get_db)):
    sale = db.query(models.Sale).filter(models.Sale.sale_id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail=f"Sale {sale_id} not found")

    sale.store_id = sale_input.store_id
    db.query(models.SaleLine).filter(models.SaleLine.sale_id == sale_id).delete()
    total_amount = 0
    sale_lines_instances = []

    for line in sale_input.lines:
        if line.quantity <= 0:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Quantity must be > 0 for batch {line.batch_id}")

        batch = db.query(models.Batch).filter(models.Batch.batch_id == line.batch_id, models.Batch.is_active.is_(True)).first()
        if not batch:
            db.rollback()
            raise HTTPException(status_code=404, detail=f"Batch {line.batch_id} not found or inactive")

        product = db.query(models.Product).filter(models.Product.product_id == batch.product_id, models.Product.is_active.is_(True)).first()
        if not product:
            db.rollback()
            raise HTTPException(status_code=404, detail=f"Product for batch {line.batch_id} not found or inactive")

        fifo_check = FIFOService.check_fifo_violation(db, sale.store_id, product.product_id, line.batch_id)
        if fifo_check['is_violation']:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"FIFO violation: expected batch {fifo_check['expected_batch_code']} for product {product.name}")

        stock_record = db.query(models.Stock).filter(models.Stock.store_id == sale.store_id, models.Stock.batch_id == line.batch_id).first()
        if not stock_record or stock_record.quantity < line.quantity:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Insufficient stock for batch {line.batch_id}")

        stock_record.quantity -= line.quantity
        subtotal = float(product.unit_price) * line.quantity
        if line.subtotal and abs(line.subtotal - subtotal) > 0.01:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Invalid subtotal for batch {line.batch_id}")

        sale_line = models.SaleLine(sale_id=sale.sale_id, batch_id=line.batch_id, quantity=line.quantity, subtotal=subtotal)
        db.add(sale_line)
        sale_lines_instances.append(sale_line)
        total_amount += subtotal

    sale.total_amount = total_amount
    db.commit()
    db.refresh(sale)

    return schemas.SaleResponse(
        sale_id=sale.sale_id,
        store_id=sale.store_id,
        total_amount=sale.total_amount,
        date=sale.date,
        lines=sale_lines_instances
    )

# Delete sale
@router.delete("/{sale_id}")
def delete_sale_fifo(sale_id: int, db: Session = Depends(database.get_db)):
    sale = db.query(models.Sale).filter(models.Sale.sale_id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail=f"Sale {sale_id} not found")
    db.query(models.SaleLine).filter(models.SaleLine.sale_id == sale_id).delete()
    db.delete(sale)
    db.commit()
    return {"message": f"Sale {sale_id} deleted successfully"}

# Daily sales report (pdf/excel)
@router.get("/daily-report")
def get_daily_sales_report(
    store_id: int,
    format: str = Query(..., description="Report format: 'pdf' or 'excel'"),
    report_date: date = None,
    db: Session = Depends(database.get_db)
):
    if report_date is None:
        report_date = date.today()
    store = StockService.get_store(db, store_id)
    start_dt = datetime.combine(report_date, datetime.min.time())
    end_dt = datetime.combine(report_date, datetime.max.time())
    sales = db.query(models.Sale).filter(models.Sale.store_id == store_id, models.Sale.date >= start_dt, models.Sale.date <= end_dt).all()

    sales_data = []
    total_amount = 0
    for sale in sales:
        lines = db.query(models.SaleLine, models.Batch, models.Product)\
            .join(models.Batch, models.Batch.batch_id == models.SaleLine.batch_id)\
            .join(models.Product, models.Product.product_id == models.Batch.product_id)\
            .filter(models.SaleLine.sale_id == sale.sale_id).all()
        for sale_line, batch, product in lines:
            sales_data.append({
                "sale_id": sale.sale_id,
                "product_name": product.name,
                "quantity": sale_line.quantity,
                "unit_price": float(product.unit_price),
                "subtotal": float(sale_line.subtotal)
            })
            total_amount += float(sale_line.subtotal)

    if format.lower() == "pdf":
        buffer = generate_sales_pdf_report(sales_data, store.name, report_date, total_amount)
        return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=daily_sales_report_{report_date}.pdf"})
    elif format.lower() == "excel":
        buffer = generate_sales_excel_report(sales_data, store.name, report_date, total_amount)
        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=daily_sales_report_{report_date}.xlsx"})
    else:
        raise HTTPException(status_code=400, detail="Format must be 'pdf' or 'excel'")

# PDF report generator
def generate_sales_pdf_report(sales_data: List[dict], store_name: str, report_date: date, total_amount: float):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a1a1a'), spaceAfter=30, alignment=1)
    elements.append(Paragraph("Daily Sales Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    info_data = [
        ['Store:', store_name],
        ['Report Date:', report_date.strftime('%Y-%m-%d')],
        ['Total Sales:', f'€{total_amount:.2f}'],
        ['Total Transactions:', str(len(sales_data))]
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([('FONTNAME', (0,0),(0,-1), 'Helvetica-Bold'), ('FONTNAME', (1,0),(1,-1), 'Helvetica'), ('FONTSIZE', (0,0), (-1,-1), 11), ('BOTTOMPADDING', (0,0), (-1,-1), 12), ('TOPPADDING', (0,0), (-1,-1), 12)]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    if sales_data:
        table_data = [['Sale ID', 'Product', 'Quantity', 'Unit Price', 'Subtotal']]
        for item in sales_data:
            table_data.append([item['sale_id'], item['product_name'], item['quantity'], f"€{item['unit_price']:.2f}", f"€{item['subtotal']:.2f}"])
        sales_table = Table(table_data, colWidths=[1*inch, 2.5*inch, 1*inch, 1.2*inch, 1.3*inch])
        sales_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,0), 12), ('BOTTOMPADDING', (0,0), (-1,0), 12), ('BACKGROUND', (0,1), (-1,-1), colors.beige), ('FONTNAME', (0,1), (-1,-1), 'Helvetica'), ('FONTSIZE', (0,1), (-1,-1), 10), ('GRID', (0,0), (-1,-1), 1, colors.black), ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey])]))
        elements.append(sales_table)
    else:
        elements.append(Paragraph("No sales recorded for this date.", styles['Normal']))
    doc.build(elements)
    buffer.seek(0)
    return buffer

# Excel report generator
def generate_sales_excel_report(sales_data: List[dict], store_name: str, report_date: date, total_amount: float):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Sales Report"
    ws['A1'] = "Daily Sales Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Store:"; ws['B3'] = store_name
    ws['A4'] = "Report Date:"; ws['B4'] = report_date.strftime('%Y-%m-%d')
    ws['A5'] = "Total Sales:"; ws['B5'] = f'€{total_amount:.2f}'
    ws['A6'] = "Total Transactions:"; ws['B6'] = len(sales_data)

    headers = ['Sale ID', 'Product', 'Quantity', 'Unit Price', 'Subtotal']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(sales_data, start=9):
        ws.cell(row=row_idx, column=1).value = item['sale_id']
        ws.cell(row=row_idx, column=2).value = item['product_name']
        ws.cell(row=row_idx, column=3).value = item['quantity']
        ws.cell(row=row_idx, column=4).value = item['unit_price']
        ws.cell(row=row_idx, column=5).value = item['subtotal']
        ws.cell(row=row_idx, column=4).number_format = '€#,##0.00'
        ws.cell(row=row_idx, column=5).number_format = '€#,##0.00'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
